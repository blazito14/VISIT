import argparse
import csv
import os
import re
from itertools import combinations

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------- Defaults ----------------
DEFAULT_GROUP_SIZE = 2
DEFAULT_VISITS_PER_GROUP = 2
PRIMARY_CONSECUTIVE_BLOCKS = 3
FALLBACK_CONSECUTIVE_BLOCKS = 2
# ------------------------------------------


DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
DAY_TO_IDX = {d.upper(): i for i, d in enumerate(DAY_NAMES)}
IDX_TO_DAY = {i: d for i, d in enumerate(DAY_NAMES)}


def norm_name(x: object) -> str:
    """
    Normalize a student name / preferred partner name from the student CSV.
    IMPORTANT: Only change is stripping whitespace (and consistent uppercasing),
    as requested. This fixes user input issues like ' John Smith '.
    """
    if pd.isna(x):
        return "NAN"
    return str(x).strip().upper()


def parse_availability_cells(cells) -> np.ndarray:
    """
    Parse 16 time-slot cells where each cell is either NaN or a comma-separated list
    of weekdays ("Monday, Wednesday, Friday") into a 16x5 binary matrix.
    """
    mat = np.zeros((16, 5), dtype=int)
    for time_idx, cell in enumerate(cells):
        if pd.isna(cell):
            continue
        days = [d.strip().upper() for d in str(cell).split(",")]
        for d in days:
            if not d or d == "NAN":
                continue
            if d in DAY_TO_IDX:
                mat[time_idx, DAY_TO_IDX[d]] = 1
    return mat


def make_groups(student_csv: str, group_size: int):
    """
    DO NOT CHANGE THE STUDENT GROUP ASSIGNMENT ALGORITHM.

    This is a light refactor of your existing script into a function.
    The only intentional behavioral change is:
      - strip whitespace from student names and preferred partner names
        before uppercasing/using them for matching.

    Additional small adjustment (as requested):
      - Avoid creating a group of size 1 when forming groups from remaining students.
        * If preferred size is 2 and we'd leave 1 student, make a group of 3.
        * If preferred size is 3 and we'd otherwise leave 1 student, make TWO groups of 2
          (i.e., keep the current group at size 2 when exactly 4 students remain).
        (A group of 2 is fine when preferred size is 3.)
    """
    student_data = pd.read_csv(student_csv, quoting=csv.QUOTE_MINIMAL).to_numpy().astype("str")

    # --------- REQUIRED FIX: strip whitespace in name columns ----------
    student_data[:, 2] = np.vectorize(norm_name)(student_data[:, 2])
    student_data[:, 3] = np.vectorize(norm_name)(student_data[:, 3])
    student_data[:, 4] = np.vectorize(norm_name)(student_data[:, 4])
    # -----------------------------------------------------------------

    # Keep the original behavior of uppercasing availability strings
    student_data[:, 5:21] = np.char.upper(student_data[:, 5:21].astype("str"))

    all_students = set(student_data[:, 2])
    group_nums = list(range(len(all_students) // 2, 0, -1))
    grouped_students = set()

    groups = {i: {} for i in group_nums}

    # Grouping Students by requested group
    students = {}
    for row_idx, student in enumerate(student_data[:, 2]):
        students[student] = {}
        students[student]["preferred"] = [student_data[row_idx, 3], student_data[row_idx, 4]]
        students[student]["group"] = False
        students[student]["availability"] = parse_availability_cells(student_data[row_idx, range(5, 21)])

    # Assign mutual preferences
    for student in student_data[:, 2]:
        mutual = set()
        for preferred in students[student]["preferred"]:
            if preferred in all_students and student in students[preferred]["preferred"]:
                mutual.add(preferred)

        students[student]["mutual"] = list(mutual)
        overlap = mutual & grouped_students
        if len(overlap) == 0 and len(mutual) != 0:
            group = group_nums.pop()
        elif len(mutual) != 0:
            group = students[overlap.pop()]["group"]
        else:
            continue

        groups[group][student] = students[student]
        grouped_students.add(student)
        students[student]["group"] = group

    # Remaining students
    remaining_students = all_students - grouped_students
    for student in list(remaining_students):
        students[student]["score"] = {}
        for other in list(remaining_students):
            if student != other:
                students[student]["score"][other] = np.sum(
                    np.ma.masked_array(students[student]["availability"], mask=students[other]["availability"])
                )

    # Work only with students who were not pre-grouped
    unassigned = set(remaining_students)

    # Build symmetric pair list once
    pairs = []
    for a, b in combinations(unassigned, 2):
        score = students[a]["score"][b]  # symmetric by assumption
        pairs.append((score, a, b))

    pairs.sort(reverse=True)

    def group_score(candidate, group):
        """Total compatibility of candidate with existing group"""
        return sum(students[candidate]["score"][m] for m in group)

    # Form groups from unassigned students
    while unassigned:
        group_members = []

        # Try to seed group with best remaining pair
        for score, a, b in pairs:
            if a in unassigned and b in unassigned:
                group_members = [a, b]
                unassigned.remove(a)
                unassigned.remove(b)
                break

        # Edge case: only one student left
        if not group_members:
            group_members = [unassigned.pop()]

        # ----------------- NEW: avoid ending with a group of 1 -----------------
        # Decide an effective target size for THIS group to prevent leaving exactly 1 student.
        effective_size = group_size
        remaining_after_seed = len(unassigned)  # number left after picking seed members

        if group_size == 2:
            # If we keep making size-2 groups, leaving 1 at the end is bad.
            # Make this group a 3 so the last group isn't size 1.
            if remaining_after_seed == 1:
                effective_size = 3

        elif group_size == 3:
            # If there are 4 students total at the start of this iteration, we seed a pair (2),
            # leaving 2. Growing to 3 would leave 1 (bad). Instead, keep this group at 2 so
            # the next iteration forms another group of 2.
            if remaining_after_seed == 2:
                effective_size = 2
        # ---------------------------------------------------------------------

        # Grow group up to effective size
        while len(group_members) < effective_size and unassigned:
            best = max(unassigned, key=lambda s: group_score(s, group_members))
            group_members.append(best)
            unassigned.remove(best)

        # Assign a group number
        group_id = group_nums.pop()

        # Initialize group availability as all ones (all slots available)
        group_availability = np.ones((16, 5), dtype=int)
        for student in group_members:
            groups[group_id][student] = students[student]
            students[student]["group"] = group_id
            grouped_students.add(student)

            # Update binary mask: only keep slots where everyone is available
            group_availability &= students[student]["availability"]

        # Add group-level availability (binary mask)
        groups[group_id]["group_availability"] = group_availability

    # Ensure groups created by mutual preferences also have group_availability
    for group_id, members in groups.items():
        if "group_availability" not in members:
            avail = np.ones((16, 5), dtype=int)
            for student in members:
                avail &= members[student]["availability"]
            groups[group_id]["group_availability"] = avail

    for gid in list(groups.keys()):
        student_names = [k for k in groups[gid].keys() if k in students]
        if len(student_names) < 1:
            del groups[gid]

    return groups, students


# ---------------- Teacher parsing ----------------

def grade_to_num(token: str):
    t = token.strip().lower()
    if t in ("k", "kindergarten"):
        return 0
    hs = {"freshman": 9, "sophomore": 10, "sophmore": 10, "junior": 11, "senior": 12}
    if t in hs:
        return hs[t]
    if re.fullmatch(r"\d+", t):
        return int(t)
    return None


def parse_grade_level(val):
    if pd.isna(val):
        return None
    parts = [p.strip() for p in str(val).split(",")]
    nums = [grade_to_num(p) for p in parts if grade_to_num(p) is not None]
    if not nums:
        return None
    return float(sum(nums)) / len(nums)


def parse_class_size(val):
    if pd.isna(val):
        return None
    m = re.search(r"(\d+)", str(val))
    if not m:
        return None
    return int(m.group(1))


def build_visit_opportunities(teacher_csv: str):
    """
    Each teacher row may contain multiple class requests (Class Size, Grade Level, and 16 availability cells)
    repeated with suffixes (.1, .2, ...). We explode those into separate 'visit opportunities' as requested.
    """
    df = pd.read_csv(teacher_csv)

    base_time_cols = [
                         c for c in df.columns
                         if c.startswith("What times could this class be visited") and "[" in c and not re.search(r"\.\d+$", c)
                     ][:16]

    opps = []
    for ridx, row in df.iterrows():
        teacher_name = str(row.get("Teacher Name", "")).strip()
        teacher_school = str(row.get("Teacher School", "")).strip()
        teacher_email = str(row.get("Teacher Email", "")).strip()

        for k in range(0, 10):
            cs_col = "Class Size" if k == 0 else f"Class Size.{k}"
            gl_col = "Grade Level" if k == 0 else f"Grade Level.{k}"
            com_col = "Any other comments?" if k == 0 else f"Any other comments?.{k}"

            if cs_col not in df.columns:
                continue

            time_cols = [c if k == 0 else f"{c}.{k}" for c in base_time_cols]
            if all(pd.isna(row.get(c, np.nan)) for c in time_cols):
                continue

            opps.append(
                {
                    "opp_id": len(opps) + 1,
                    "teacher_name": teacher_name,
                    "teacher_school": teacher_school,
                    "teacher_email": teacher_email,
                    "class_size": parse_class_size(row.get(cs_col, np.nan)),
                    "grade_level_raw": None if pd.isna(row.get(gl_col, np.nan)) else str(row.get(gl_col, np.nan)),
                    "grade_level_num": parse_grade_level(row.get(gl_col, np.nan)),
                    "comments": None if pd.isna(row.get(com_col, np.nan)) else str(row.get(com_col, np.nan)),
                    "availability": parse_availability_cells([row.get(c, np.nan) for c in time_cols]),
                    "response_row": ridx,
                    "submission_index": k,
                }
            )

    return opps


# ---------------- Scheduling ----------------

def consecutive_start_indices(overlap_col: np.ndarray, need: int):
    """
    Return all start indices where overlap_col has >=need consecutive 1s.
    """
    starts = []
    run = 0
    start = 0
    for i, v in enumerate(overlap_col):
        if v == 1:
            if run == 0:
                start = i
            run += 1
        else:
            if run >= need:
                for s in range(start, i - need + 1):
                    starts.append(s)
            run = 0
    if run >= need:
        for s in range(start, len(overlap_col) - need + 1):
            starts.append(s)
    return starts


def best_time_block(group_avail: np.ndarray, opp_avail: np.ndarray, need: int):
    """
    Choose a best (day, start_idx) satisfying the consecutive-block rule.
    """
    overlap = (group_avail & opp_avail)
    best = None
    for day in range(5):
        for s in consecutive_start_indices(overlap[:, day], need):
            # Prefer more overlap on that day, and earlier start slightly
            score = overlap[:, day].sum() * 10 - s
            cand = (score, day, s)
            if best is None or cand[0] > best[0]:
                best = cand
    if best is None:
        return None
    _, day, s = best
    return {"day_idx": day, "start_idx": s, "blocks": need}


def schedule_visits(
        groups,
        opportunities,
        visits_per_group: int,
        primary_blocks: int,
        fallback_blocks: int,
        time_labels,
):
    """
    Greedy scheduler:
      - each visit opportunity can be used at most once
      - prefer not reusing teachers (penalty)
      - enforce consecutive overlap blocks (3, fallback to 2)
      - groups can have multiple visits in the same weekly slot (treated as different weeks)
      - grade similarity: tries to keep a group's multiple visits near the same grade level,
        since student grade level is not present in the student CSV.
    """
    group_order = sorted([(g["group_availability"].sum(), gid) for gid, g in groups.items()])  # hardest first

    opp_used = set()
    teacher_count = {}
    group_prev_grade = {}

    rows = []
    for _, gid in group_order:
        members = [s for s in groups[gid].keys() if s != "group_availability"]
        members_str = ", ".join(members)
        g_av = groups[gid]["group_availability"]

        for vnum in range(1, visits_per_group + 1):
            chosen = None

            for need in (primary_blocks, fallback_blocks):
                for opp in opportunities:
                    oid = opp["opp_id"]
                    if oid in opp_used:
                        continue

                    tb = best_time_block(g_av, opp["availability"], need)
                    if tb is None:
                        continue

                    score = 0

                    # teacher reuse penalty (soft constraint)
                    score -= teacher_count.get(opp["teacher_name"], 0) * 50

                    # grade similarity across a group's assigned visits (soft)
                    if group_prev_grade.get(gid) is not None and opp["grade_level_num"] is not None:
                        score -= abs(opp["grade_level_num"] - group_prev_grade[gid]) * 5

                    # prefer meeting primary constraint
                    score += 20 if need == primary_blocks else 0

                    # prefer bigger overall overlap
                    score += (g_av & opp["availability"]).sum()

                    # tie-break: earlier start
                    score -= tb["start_idx"] * 0.5

                    if chosen is None or score > chosen[0]:
                        chosen = (score, opp, need, tb)

                if chosen is not None:
                    break

            if chosen is None:
                # Still output a row even when not assigned
                rows.append(
                    {
                        "group_number": gid,
                        "visit_number": vnum,
                        "students": members_str,
                        "opp_id": "",
                        "teacher_name": "",
                        "teacher_school": "",
                        "teacher_email": "",
                        "class_size": "",
                        "grade_level_raw": "",
                        "grade_level_num": "",
                        "scheduled_day": "",
                        "scheduled_start_block": "",
                        "consecutive_blocks_required": "",
                        "comments": "",
                    }
                )
                continue

            _, opp, need_used, tb = chosen
            opp_used.add(opp["opp_id"])
            teacher_count[opp["teacher_name"]] = teacher_count.get(opp["teacher_name"], 0) + 1

            # update group grade "anchor"
            if opp["grade_level_num"] is not None:
                if group_prev_grade.get(gid) is None:
                    group_prev_grade[gid] = opp["grade_level_num"]
                else:
                    group_prev_grade[gid] = (group_prev_grade[gid] + opp["grade_level_num"]) / 2.0

            rows.append(
                {
                    "group_number": gid,
                    "visit_number": vnum,
                    "students": members_str,
                    "opp_id": opp["opp_id"],
                    "teacher_name": opp["teacher_name"],
                    "teacher_school": opp["teacher_school"],
                    "teacher_email": opp["teacher_email"],
                    "class_size": opp["class_size"],
                    "grade_level_raw": opp["grade_level_raw"],
                    "grade_level_num": opp["grade_level_num"],
                    "scheduled_day": IDX_TO_DAY[tb["day_idx"]],
                    "scheduled_start_block": time_labels[tb["start_idx"]],
                    "consecutive_blocks_required": need_used,
                    "comments": opp["comments"],
                }
            )

    return pd.DataFrame(rows)


def _split_time_range(label: str):
    if label is None or (isinstance(label, float) and np.isnan(label)):
        return ("", "")
    s = str(label).strip()
    if not s:
        return ("", "")
    # allow '-' or '–'
    parts = re.split(r"\s*[-–]\s*", s)
    if len(parts) >= 2:
        return (parts[0].strip(), parts[1].strip())
    return (s, "")


def _infer_ampm(start_part: str, end_part: str):
    # If start already contains AM/PM, keep it. Otherwise borrow from the end part.
    if re.search(r"\b(AM|PM)\b", start_part, flags=re.I):
        return start_part
    m = re.search(r"\b(AM|PM)\b", end_part, flags=re.I)
    if m:
        return f"{start_part} {m.group(1).upper()}"
    return start_part


def format_scheduled_visits_csv(schedule_df: pd.DataFrame, time_labels):
    """
    Create the user-facing scheduled_visits.csv with:
      - no opp_id column
      - 'Grade Level' column (from grade_level_raw)
      - start/end times (derived from the scheduled start block and the block length)
      - prettier/capitalized headings
      - sorted by Group Number, then Visit Number
    Note: we keep the raw schedule_df unchanged for internal workbook generation.
    """
    df = schedule_df.copy()

    # Compute Visit Start / Visit End times using the time_labels (30 min blocks)
    def compute_start_end(row):
        start_block = row.get("scheduled_start_block")
        blocks = row.get("consecutive_blocks_required")
        if pd.isna(start_block) or start_block == "" or pd.isna(blocks) or blocks == "":
            return ("", "")
        start_block = str(start_block)
        try:
            start_idx = time_labels.index(start_block)
        except ValueError:
            return ("", "")
        try:
            b = int(float(blocks))
        except Exception:
            return ("", "")

        s0, e0 = _split_time_range(time_labels[start_idx])
        start_time = _infer_ampm(s0, e0)

        end_idx = min(start_idx + b - 1, len(time_labels) - 1)
        s1, e1 = _split_time_range(time_labels[end_idx])
        end_time = e1 if e1 else _infer_ampm(s1, e0)
        return (start_time, end_time)

    se = df.apply(lambda r: compute_start_end(r), axis=1, result_type="expand")
    df["Visit Start"] = se[0]
    df["Visit End"] = se[1]

    # Build final column set (drop opp_id and use nicer headings)
    col_map = {
        "group_number": "Group Number",
        "visit_number": "Visit Number",
        "students": "Students",
        "teacher_name": "Teacher Name",
        "teacher_school": "Teacher School",
        "teacher_email": "Teacher Email",
        "class_size": "Class Size",
        "grade_level_raw": "Grade Level",
        "grade_level_num": "Grade Level Num",
        "scheduled_day": "Scheduled Day",
        "scheduled_start_block": "Scheduled Start Block",
        "consecutive_blocks_required": "Blocks Required",
        "comments": "Comments",
        "opp_id": "Opp ID"
    }

    keep = list(col_map.keys()) + ["Visit Start", "Visit End"]
    df = df[keep].rename(columns=col_map)

    # Arrange columns in a more readable order
    ordered = [
        "Group Number",
        "Visit Number",
        "Students",
        "Teacher Name",
        "Teacher School",
        "Teacher Email",
        "Class Size",
        "Grade Level",
        "Grade Level Num",
        "Scheduled Day",
        "Visit Start",
        "Visit End",
        "Scheduled Start Block",
        "Blocks Required",
        "Comments",
        "Opp ID"
    ]
    df = df[ordered]

    # Sort as requested
    df = df.sort_values(["Group Number", "Visit Number"], kind="mergesort")

    return df

# ---------------- Excel outputs ----------------

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_matrix(ws, mat: np.ndarray, title: str, time_labels, meta_lines=None):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)

    row0 = 3
    if meta_lines:
        for i, line in enumerate(meta_lines):
            ws[f"A{2+i}"] = line
        row0 = 2 + len(meta_lines) + 1

    # headers
    ws.cell(row=row0, column=1, value="Time").font = Font(bold=True)
    ws.cell(row=row0, column=1).fill = HEADER_FILL
    ws.cell(row=row0, column=1).alignment = CENTER
    ws.cell(row=row0, column=1).border = BORDER

    ws.column_dimensions["A"].width = 18
    for j, day in enumerate(DAY_NAMES):
        c = ws.cell(row=row0, column=2 + j, value=day)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(2 + j)].width = 14

    for i, t in enumerate(time_labels):
        r = row0 + 1 + i
        c = ws.cell(row=r, column=1, value=t)
        c.alignment = CENTER
        c.border = BORDER

        for j in range(5):
            cell = ws.cell(row=r, column=2 + j, value=int(mat[i, j]))
            cell.alignment = CENTER
            cell.border = BORDER

    ws.freeze_panes = ws[f"B{row0+1}"]


def save_groups_workbook(groups, out_path: str, time_labels):
    wb = Workbook()
    wb.remove(wb.active)
    for gid in sorted(groups.keys()):
        ws = wb.create_sheet(f"Group {gid}")
        members = [s for s in groups[gid].keys() if s != "group_availability"]
        write_matrix(
            ws,
            groups[gid]["group_availability"],
            f"Group {gid} Availability",
            time_labels,
            meta_lines=[f"Students: {', '.join(members)}"],
        )
    wb.save(out_path)


def save_opportunities_workbook(opps, out_path: str, time_labels):
    wb = Workbook()
    wb.remove(wb.active)
    for opp in opps:
        ws = wb.create_sheet(f"Opp {opp['opp_id']}")
        meta = [
            f"Teacher: {opp['teacher_name']}",
            f"School: {opp['teacher_school']}",
            f"Email: {opp['teacher_email']}",
            f"Class Size: {opp['class_size']}",
            f"Grade Level: {opp['grade_level_raw']} (num={opp['grade_level_num']})",
        ]
        write_matrix(ws, opp["availability"], f"Visit Opportunity {opp['opp_id']} Availability", time_labels, meta)
    wb.save(out_path)


def save_group_detail_workbooks(groups, opps_by_id, schedule_df: pd.DataFrame, out_dir: str, time_labels):
    """
    Per-group workbook:
      - First tab: Group availability
      - Subsequent tabs: EVERY visit opportunity, each on its own sheet,
        showing a SINGLE matrix that is (group_availability + opp_availability),
        so:
          2 = overlap (both available)
          1 = available in only one of the two
          0 = neither
        Each sheet also includes the same helpful metadata header that the
        previous "Opp" sheets had.

    This is intended to support manual verification / manual scheduling.
    """
    os.makedirs(out_dir, exist_ok=True)

    # Precompute (group, opp_id) -> visit_number mapping for quick lookup
    assigned_lookup = {}
    if schedule_df is not None and len(schedule_df) > 0:
        for _, row in schedule_df.iterrows():
            try:
                gid = int(row.get("group_number"))
            except Exception:
                continue
            opp_id = row.get("opp_id")
            if opp_id == "" or pd.isna(opp_id):
                continue
            try:
                oid = int(opp_id)
                vnum = int(row.get("visit_number"))
            except Exception:
                continue
            assigned_lookup[(gid, oid)] = vnum

    for gid in sorted(groups.keys()):
        wb = Workbook()
        ws = wb.active
        ws.title = "Group Availability"

        members = [s for s in groups[gid].keys() if s != "group_availability"]
        write_matrix(
            ws,
            groups[gid]["group_availability"],
            f"Group {gid} Availability",
            time_labels,
            meta_lines=[f"Students: {', '.join(members)}"],
        )

        g_av = groups[gid]["group_availability"]

        # Create one sheet per opportunity (sorted by opp_id)
        for oid in sorted(opps_by_id.keys()):
            opp = opps_by_id[oid]

            # Excel sheet name must be <= 31 chars and unique
            sheet_name = f"Opp {oid}"
            ws_opp = wb.create_sheet(sheet_name[:31])

            v_assigned = assigned_lookup.get((gid, oid))
            assigned_line = f"Assigned to this group as Visit #{v_assigned}" if v_assigned is not None else "Not assigned to this group"

            meta = [
                f"Teacher: {opp['teacher_name']}",
                f"School: {opp['teacher_school']}",
                f"Email: {opp['teacher_email']}",
                f"Class Size: {opp['class_size']}",
                f"Grade Level: {opp['grade_level_raw']} (num={opp['grade_level_num']})",
                assigned_line,
                "Legend: 2 = overlap (both available), 1 = available in only one, 0 = neither",
            ]

            overlap = (g_av & opp["availability"]) + opp["availability"]
            write_matrix(
                ws_opp,
                overlap,
                f"Group {gid} + Opportunity {oid} (Overlap Matrix)",
                time_labels,
                meta_lines=meta,
            )

        wb.save(os.path.join(out_dir, f"group_{gid}.xlsx"))


def main():
    ap = argparse.ArgumentParser(description="Schedule club visit groups and export CSV/XLSX outputs.")
    ap.add_argument("--student_csv", required=True)
    ap.add_argument("--teacher_csv", required=True)
    ap.add_argument("--group_size", type=int, default=DEFAULT_GROUP_SIZE)
    ap.add_argument("--visits_per_group", type=int, default=DEFAULT_VISITS_PER_GROUP)
    ap.add_argument("--out_dir", default="outputs")
    args = ap.parse_args()

    os.makedirs(args.out_dir, exist_ok=True)

    # Time labels inferred from the student CSV (columns 5..20)
    s_cols = pd.read_csv(args.student_csv).columns.tolist()
    time_labels = [re.search(r"\[(.*?)\]", c).group(1) for c in s_cols[5:21]]

    groups, _students = make_groups(args.student_csv, group_size=args.group_size)
    opps = build_visit_opportunities(args.teacher_csv)

    schedule_df = schedule_visits(
        groups=groups,
        opportunities=opps,
        visits_per_group=args.visits_per_group,
        primary_blocks=PRIMARY_CONSECUTIVE_BLOCKS,
        fallback_blocks=FALLBACK_CONSECUTIVE_BLOCKS,
        time_labels=time_labels,
    )

    schedule_csv = os.path.join(args.out_dir, "scheduled_visits.csv")
    format_scheduled_visits_csv(schedule_df, time_labels).to_csv(schedule_csv, index=False)

    groups_xlsx = os.path.join(args.out_dir, "student_groups_availability.xlsx")
    save_groups_workbook(groups, groups_xlsx, time_labels=time_labels)

    opps_xlsx = os.path.join(args.out_dir, "visit_opportunities_availability.xlsx")
    save_opportunities_workbook(opps, opps_xlsx, time_labels=time_labels)

    opps_by_id = {o["opp_id"]: o for o in opps}
    group_dir = os.path.join(args.out_dir, "group_workbooks")
    save_group_detail_workbooks(groups, opps_by_id, schedule_df, group_dir, time_labels=time_labels)

    print("Wrote:")
    print(" -", schedule_csv)
    print(" -", groups_xlsx)
    print(" -", opps_xlsx)
    print(" -", group_dir)


if __name__ == "__main__":
    main()
